"""
Excel Storage
Stores event data in Excel files using openpyxl
"""

from typing import List
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from src.storage.base_storage import BaseStorage
from src.models.event import Event
from src.utils.config import config
from src.utils.logger import setup_logger
from src.utils.helpers import is_date_expired


logger = setup_logger(__name__, config.LOG_FILE, config.LOG_LEVEL)


class ExcelStorage(BaseStorage):
    """Excel file storage backend"""
    
    def __init__(self, file_path: str = None):
        """
        Initialize Excel storage
        
        Args:
            file_path: Path to Excel file
        """
        super().__init__()
        self.file_path = file_path or config.EXCEL_FILE_PATH
        self.file_path = Path(self.file_path)
        
        # Ensure directory exists
        self.file_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Initialize file if it doesn't exist
        if not self.file_path.exists():
            self._create_new_file()
    
    def _create_new_file(self):
        """Create new Excel file with headers"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Events"
            
            # Headers
            headers = [
                'Event ID', 'Event Name', 'Date', 'Venue', 'City',
                'Category', 'URL', 'Source', 'Status', 'Last Updated'
            ]
            ws.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Set column widths
            column_widths = [15, 30, 15, 25, 15, 15, 50, 15, 12, 20]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + i)].width = width
            
            wb.save(self.file_path)
            logger.info(f"Created new Excel file: {self.file_path}")
            
        except Exception as e:
            logger.error(f"Error creating Excel file: {e}")
            raise
    
    def _headers(self):
        return [
            'Event ID', 'Event Name', 'Date', 'Venue', 'City',
            'Category', 'URL', 'Source', 'Status', 'Last Updated'
        ]

    def save_events(self, events: List[Event]) -> bool:
        """
        Save events to Excel file
        
        Args:
            events: List of Event objects
        
        Returns:
            True if successful, False otherwise
        """
        try:
            # Load existing events
            existing_events = self.load_events()
            
            # Merge with new events
            merged_events = self.merge_events(events, existing_events)
            
            # Write to Excel using openpyxl only
            wb = Workbook()
            ws = wb.active
            ws.title = "Events"

            headers = self._headers()
            ws.append(headers)

            for event in merged_events:
                data = event.to_dict()
                ws.append([
                    data['Event ID'],
                    data['Event Name'],
                    data['Date'],
                    data['Venue'],
                    data['City'],
                    data['Category'],
                    data['URL'],
                    data['Source'],
                    data['Status'],
                    data['Last Updated'],
                ])

            # Style headers
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            wb.save(self.file_path)
            
            logger.info(f"Saved {len(merged_events)} events to {self.file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error saving events to Excel: {e}")
            return False
    
    def load_events(self) -> List[Event]:
        """
        Load events from Excel file
        
        Returns:
            List of Event objects
        """
        try:
            if not self.file_path.exists():
                logger.warning(f"File not found: {self.file_path}")
                return []
            
            wb = load_workbook(self.file_path)
            if "Events" in wb.sheetnames:
                ws = wb["Events"]
            else:
                ws = wb.active

            rows = list(ws.iter_rows(values_only=True))
            if not rows or len(rows) < 2:
                return []

            headers = [str(h) for h in rows[0]]
            events = []
            for row in rows[1:]:
                if all(cell is None for cell in row):
                    continue
                row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
                try:
                    event = Event(
                        event_name=str(row_dict.get('Event Name', '')),
                        date=str(row_dict.get('Date', '')),
                        venue=str(row_dict.get('Venue', '')),
                        city=str(row_dict.get('City', '')),
                        category=str(row_dict.get('Category', '')),
                        url=str(row_dict.get('URL', '')),
                        source=str(row_dict.get('Source', '')),
                        status=str(row_dict.get('Status', '')),
                        event_id=str(row_dict.get('Event ID', ''))
                    )
                    events.append(event)
                except Exception as e:
                    logger.debug(f"Error parsing event row: {e}")
                    continue
            
            logger.info(f"Loaded {len(events)} events from {self.file_path}")
            return events
            
        except Exception as e:
            logger.error(f"Error loading events from Excel: {e}")
            return []
    
    def update_event(self, event: Event) -> bool:
        """
        Update existing event
        
        Args:
            event: Event object to update
        
        Returns:
            True if successful, False otherwise
        """
        try:
            events = self.load_events()
            
            # Find and update event
            for i, existing_event in enumerate(events):
                if existing_event.event_id == event.event_id:
                    events[i] = event
                    return self.save_events(events)
            
            # Event not found, add it
            events.append(event)
            return self.save_events(events)
            
        except Exception as e:
            logger.error(f"Error updating event: {e}")
            return False
    
    def delete_event(self, event_id: str) -> bool:
        """
        Delete event by ID
        
        Args:
            event_id: Event ID to delete
        
        Returns:
            True if successful, False otherwise
        """
        try:
            events = self.load_events()
            events = [e for e in events if e.event_id != event_id]
            return self.save_events(events)
            
        except Exception as e:
            logger.error(f"Error deleting event: {e}")
            return False
    
    def mark_expired_events(self) -> int:
        """
        Mark expired events
        
        Returns:
            Number of events marked as expired
        """
        try:
            events = self.load_events()
            expired_count = 0
            
            for event in events:
                if event.status == "Active" and is_date_expired(event.date, config.MARK_EXPIRED_DAYS):
                    event.status = "Expired"
                    event.last_updated = datetime.now()
                    expired_count += 1
            
            if expired_count > 0:
                self.save_events(events)
                logger.info(f"Marked {expired_count} events as expired")
            
            return expired_count
            
        except Exception as e:
            logger.error(f"Error marking expired events: {e}")
            return 0
